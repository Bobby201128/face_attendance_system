# -*- coding: utf-8 -*-
"""
人脸识别签到系统 - 数据库管理层
SQLite数据库，管理人员信息、签到记录、系统配置等
"""
import sqlite3
import os
import json
import logging
from datetime import datetime, date, timedelta
from contextlib import contextmanager

import config

logger = logging.getLogger(__name__)


class Database:
    """数据库管理类"""

    def __init__(self, db_path=None):
        self.db_path = db_path or config.DATABASE_PATH
        self._init_db()

    @contextmanager
    def get_connection(self):
        """获取数据库连接 (上下文管理器)"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
        try:
            yield conn
            conn.commit()
        except Exception as e:
            conn.rollback()
            logger.error(f"数据库操作失败: {e}")
            raise
        finally:
            conn.close()

    def _init_db(self):
        """初始化数据库表结构"""
        with self.get_connection() as conn:
            cursor = conn.cursor()

            # 人员信息表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS persons (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    employee_id TEXT UNIQUE,
                    department TEXT DEFAULT '',
                    position TEXT DEFAULT '',
                    phone TEXT DEFAULT '',
                    email TEXT DEFAULT '',
                    face_encoding BLOB,
                    face_image_path TEXT,
                    status INTEGER DEFAULT 1,
                    remark TEXT DEFAULT '',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # 签到记录表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS attendance (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    person_id INTEGER NOT NULL,
                    sign_type TEXT NOT NULL CHECK(sign_type IN ('in', 'out')),
                    sign_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    confidence REAL,
                    camera_id INTEGER DEFAULT 0,
                    remark TEXT DEFAULT '',
                    FOREIGN KEY (person_id) REFERENCES persons(id)
                )
            """)

            # 系统配置表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS settings (
                    key TEXT PRIMARY KEY,
                    value TEXT,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # 操作日志表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS operation_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    operator TEXT DEFAULT 'system',
                    action TEXT NOT NULL,
                    detail TEXT DEFAULT '',
                    ip_address TEXT DEFAULT '',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # 环境配置表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS environments (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    description TEXT DEFAULT '',
                    work_start_hour INTEGER DEFAULT 9,
                    work_start_minute INTEGER DEFAULT 0,
                    work_end_hour INTEGER DEFAULT 18,
                    work_end_minute INTEGER DEFAULT 0,
                    late_grace_minutes INTEGER DEFAULT 15,
                    recognition_threshold REAL DEFAULT 0.55,
                    confirm_frames INTEGER DEFAULT 3,
                    sign_cooldown_seconds INTEGER DEFAULT 60,
                    max_sign_count INTEGER DEFAULT 0,
                    sound_enabled INTEGER DEFAULT 0,
                    sound_volume REAL DEFAULT 0.8,
                    sound_text TEXT DEFAULT '',
                    sound_read_name INTEGER DEFAULT 1,
                    sign_in_required INTEGER DEFAULT 1,
                    sign_out_required INTEGER DEFAULT 1,
                    sign_mode TEXT DEFAULT 'auto',
                    is_active INTEGER DEFAULT 1,
                    default_env INTEGER DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # 创建索引
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_person ON attendance(person_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_time ON attendance(sign_time)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance(date(sign_time))")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_persons_employee ON persons(employee_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_persons_status ON persons(status)")

            # 初始化默认配置
            self._init_default_settings(cursor)

            # 初始化默认环境（如果没有）
            cursor.execute("SELECT COUNT(*) as cnt FROM environments")
            if cursor.fetchone()['cnt'] == 0:
                cursor.execute("""
                    INSERT INTO environments (name, description, default_env, is_active)
                    VALUES ('默认环境', '默认签到环境', 1, 1)
                """)

            # 确保环境表有 max_sign_count 列
            try:
                cursor.execute("ALTER TABLE environments ADD COLUMN max_sign_count INTEGER DEFAULT 0")
            except Exception:
                pass
            try:
                cursor.execute("ALTER TABLE environments ADD COLUMN sound_enabled INTEGER DEFAULT 0")
            except Exception:
                pass
            try:
                cursor.execute("ALTER TABLE environments ADD COLUMN sound_volume REAL DEFAULT 0.8")
            except Exception:
                pass
            try:
                cursor.execute("ALTER TABLE environments ADD COLUMN sound_text TEXT DEFAULT ''")
            except Exception:
                pass
            try:
                cursor.execute("ALTER TABLE environments ADD COLUMN sound_read_name INTEGER DEFAULT 1")
            except Exception:
                pass

    def _init_default_settings(self, cursor):
        """初始化默认系统配置"""
        defaults = {
            "work_start": "09:00",
            "work_end": "18:00",
            "late_grace": "15",
            "sign_mode": "auto",
            "recognition_threshold": "0.6",
            "confirm_frames": "3",
            "sign_cooldown": "60",
            "camera_index": "0",
            "system_name": "人脸识别签到系统",
            "admin_password": "admin123",
        }
        for key, value in defaults.items():
            cursor.execute(
                "INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)",
                (key, value)
            )

    # ==================== 人员管理 ====================

    def add_person(self, name, phone="", face_encoding=None, face_image_path="", remark=""):
        """添加人员 - 简化版：只保留姓名和手机号（备注）"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO persons (name, phone, face_encoding, face_image_path, remark)
                VALUES (?, ?, ?, ?, ?)
            """, (name, phone, face_encoding, face_image_path, remark))
            return cursor.lastrowid

    def update_person(self, person_id, **kwargs):
        """更新人员信息 - 简化版"""
        allowed_fields = ['name', 'phone', 'face_encoding', 'face_image_path', 'remark']
        updates = []
        values = []
        for field, value in kwargs.items():
            if field in allowed_fields:
                updates.append(f"{field} = ?")
                values.append(value)

        if not updates:
            return False

        updates.append("updated_at = CURRENT_TIMESTAMP")
        values.append(person_id)

        with self.get_connection() as conn:
            conn.execute(
                f"UPDATE persons SET {', '.join(updates)} WHERE id = ?",
                values
            )
            return True

    def delete_person(self, person_id):
        """删除人员 (软删除)"""
        with self.get_connection() as conn:
            conn.execute("UPDATE persons SET status = 0 WHERE id = ?", (person_id,))
            return True

    def hard_delete_person(self, person_id):
        """彻底删除人员"""
        with self.get_connection() as conn:
            conn.execute("DELETE FROM attendance WHERE person_id = ?", (person_id,))
            conn.execute("DELETE FROM persons WHERE id = ?", (person_id,))
            return True

    def get_person(self, person_id):
        """获取单个人员信息"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM persons WHERE id = ?", (person_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_person_by_employee_id(self, employee_id):
        """根据工号获取人员"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM persons WHERE employee_id = ? AND status = 1",
                          (employee_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_all_persons(self, include_inactive=False, search="", department="",
                       page=1, per_page=50):
        """获取人员列表 (分页+搜索)"""
        conditions = []
        params = []

        if not include_inactive:
            conditions.append("status = 1")

        if search:
            conditions.append("(name LIKE ? OR employee_id LIKE ? OR phone LIKE ?)")
            search_pattern = f"%{search}%"
            params.extend([search_pattern, search_pattern, search_pattern])

        if department:
            conditions.append("department = ?")
            params.append(department)

        where_clause = " AND ".join(conditions) if conditions else "1=1"
        offset = (page - 1) * per_page

        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                f"SELECT * FROM persons WHERE {where_clause} ORDER BY created_at DESC LIMIT ? OFFSET ?",
                params + [per_page, offset]
            )
            persons = [dict(row) for row in cursor.fetchall()]

            cursor.execute(
                f"SELECT COUNT(*) as total FROM persons WHERE {where_clause}",
                params
            )
            total = cursor.fetchone()['total']

        return persons, total

    def get_persons_with_encoding(self):
        """获取所有有面部编码的活跃人员"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT id, name, employee_id, department, face_encoding, face_image_path "
                "FROM persons WHERE face_encoding IS NOT NULL AND status = 1"
            )
            return [dict(row) for row in cursor.fetchall()]

    def get_departments(self):
        """获取所有部门列表"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                "SELECT DISTINCT department FROM persons WHERE status = 1 AND department != '' ORDER BY department"
            )
            return [row['department'] for row in cursor.fetchall()]

    def get_person_count(self):
        """获取活跃人员总数"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) as cnt FROM persons WHERE status = 1")
            return cursor.fetchone()['cnt']

    # ==================== 签到管理 ====================

    def add_attendance(self, person_id, sign_type, confidence=None, remark=""):
        """添加签到记录"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO attendance (person_id, sign_type, confidence, remark)
                VALUES (?, ?, ?, ?)
            """, (person_id, sign_type, confidence, remark))
            return cursor.lastrowid

    def get_today_attendance(self, target_date=None):
        """获取某天的签到记录"""
        target_date = target_date or date.today().isoformat()
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT a.*, p.name, p.employee_id, p.department, p.position
                FROM attendance a
                JOIN persons p ON a.person_id = p.id
                WHERE date(a.sign_time) = ?
                ORDER BY a.sign_time DESC
            """, (target_date,))
            return [dict(row) for row in cursor.fetchall()]

    def get_person_today_status(self, person_id, target_date=None):
        """获取某人某天的签到状态"""
        target_date = target_date or date.today().isoformat()
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT sign_type, sign_time, confidence
                FROM attendance
                WHERE person_id = ? AND date(sign_time) = ?
                ORDER BY sign_time
            """, (person_id, target_date))
            records = [dict(row) for row in cursor.fetchall()]

            status = {
                "signed_in": False,
                "signed_out": False,
                "sign_in_time": None,
                "sign_out_time": None,
                "records": records
            }
            for r in records:
                if r['sign_type'] == 'in' and not status['signed_in']:
                    status['signed_in'] = True
                    status['sign_in_time'] = r['sign_time']
                elif r['sign_type'] == 'out':
                    status['signed_out'] = True
                    status['sign_out_time'] = r['sign_time']

            return status

    def get_attendance_by_date_range(self, start_date, end_date, department="",
                                     person_id=None, page=1, per_page=50):
        """按日期范围查询签到记录"""
        conditions = ["date(a.sign_time) BETWEEN ? AND ?"]
        params = [start_date, end_date]

        if department:
            conditions.append("p.department = ?")
            params.append(department)

        if person_id:
            conditions.append("a.person_id = ?")
            params.append(person_id)

        where_clause = " AND ".join(conditions)
        offset = (page - 1) * per_page

        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                f"""SELECT a.*, p.name, p.employee_id, p.department, p.position
                    FROM attendance a
                    JOIN persons p ON a.person_id = p.id
                    WHERE {where_clause}
                    ORDER BY a.sign_time DESC
                    LIMIT ? OFFSET ?""",
                params + [per_page, offset]
            )
            records = [dict(row) for row in cursor.fetchall()]

            cursor.execute(
                f"""SELECT COUNT(*) as total
                    FROM attendance a
                    JOIN persons p ON a.person_id = p.id
                    WHERE {where_clause}""",
                params
            )
            total = cursor.fetchone()['total']

        return records, total

    def get_today_statistics(self):
        """获取今日签到统计"""
        today = date.today().isoformat()
        with self.get_connection() as conn:
            cursor = conn.cursor()

            # 总人数
            cursor.execute("SELECT COUNT(*) as cnt FROM persons WHERE status = 1")
            total_persons = cursor.fetchone()['cnt']

            # 今日签到人数 (签到的不同人数)
            cursor.execute("""
                SELECT COUNT(DISTINCT person_id) as cnt
                FROM attendance
                WHERE date(sign_time) = ? AND sign_type = 'in'
            """, (today,))
            signed_in = cursor.fetchone()['cnt']

            # 今日签到总次数
            cursor.execute("""
                SELECT COUNT(*) as cnt FROM attendance WHERE date(sign_time) = ?
            """, (today,))
            total_records = cursor.fetchone()['cnt']

            # 迟到人数
            settings = self.get_settings()
            work_start = settings.get('work_start', '09:00') or '09:00'
            late_grace = int(settings.get('late_grace', '15') or 15)

            try:
                cursor.execute("""
                    SELECT COUNT(DISTINCT person_id) as cnt
                    FROM attendance
                    WHERE date(sign_time) = ? AND sign_type = 'in'
                      AND time(sign_time) > time(?)
                """, (today, f"{work_start}:{late_grace:02d}:00"))
                late_count = cursor.fetchone()['cnt']
            except Exception:
                late_count = 0

            return {
                "total_persons": total_persons,
                "signed_in": signed_in,
                "absent": total_persons - signed_in,
                "late_count": late_count,
                "total_records": total_records,
                "date": today,
                "sign_rate": round(signed_in / total_persons * 100, 1) if total_persons > 0 else 0
            }

    def get_monthly_statistics(self, year=None, month=None):
        """获取月度统计"""
        today = date.today()
        year = year or today.year
        month = month or today.month

        start_date = date(year, month, 1).isoformat()
        if month == 12:
            end_date = date(year + 1, 1, 1).isoformat()
        else:
            end_date = date(year, month + 1, 1).isoformat()

        with self.get_connection() as conn:
            cursor = conn.cursor()

            # 每日签到人数统计
            cursor.execute("""
                SELECT date(sign_time) as day,
                       COUNT(DISTINCT CASE WHEN sign_type='in' THEN person_id END) as sign_in_count,
                       COUNT(*) as total_records
                FROM attendance
                WHERE date(sign_time) >= ? AND date(sign_time) < ?
                GROUP BY date(sign_time)
                ORDER BY day
            """, (start_date, end_date))
            daily_stats = [dict(row) for row in cursor.fetchall()]

            # 月度汇总
            cursor.execute("SELECT COUNT(*) as cnt FROM persons WHERE status = 1")
            total_persons = cursor.fetchone()['cnt']

            cursor.execute("""
                SELECT COUNT(DISTINCT person_id) as cnt
                FROM attendance
                WHERE date(sign_time) >= ? AND date(sign_time) < ? AND sign_type = 'in'
            """, (start_date, end_date))
            unique_signed = cursor.fetchone()['cnt']

            return {
                "year": year,
                "month": month,
                "total_persons": total_persons,
                "unique_signed": unique_signed,
                "daily_stats": daily_stats
            }

    def get_person_attendance_summary(self, person_id, year=None, month=None):
        """获取个人签到汇总"""
        today = date.today()
        year = year or today.year
        month = month or today.month

        start_date = date(year, month, 1).isoformat()
        if month == 12:
            end_date = date(year + 1, 1, 1).isoformat()
        else:
            end_date = date(year, month + 1, 1).isoformat()

        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT date(sign_time) as day,
                       MIN(CASE WHEN sign_type='in' THEN time(sign_time) END) as first_in,
                       MAX(CASE WHEN sign_type='out' THEN time(sign_time) END) as last_out,
                       COUNT(*) as record_count
                FROM attendance
                WHERE person_id = ? AND date(sign_time) >= ? AND date(sign_time) < ?
                GROUP BY date(sign_time)
                ORDER BY day
            """, (person_id, start_date, end_date))
            return [dict(row) for row in cursor.fetchall()]

    # ==================== 系统设置 ====================

    def get_settings(self):
        """获取所有设置"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT key, value FROM settings")
            return {row['key']: row['value'] for row in cursor.fetchall()}

    def get_setting(self, key, default=None):
        """获取单个设置"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT value FROM settings WHERE key = ?", (key,))
            row = cursor.fetchone()
            return row['value'] if row else default

    def update_setting(self, key, value):
        """更新设置"""
        with self.get_connection() as conn:
            conn.execute("""
                INSERT INTO settings (key, value, updated_at)
                VALUES (?, ?, CURRENT_TIMESTAMP)
                ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = CURRENT_TIMESTAMP
            """, (key, str(value)))

    def update_settings_batch(self, settings_dict):
        """批量更新设置"""
        with self.get_connection() as conn:
            for key, value in settings_dict.items():
                conn.execute("""
                    INSERT INTO settings (key, value, updated_at)
                    VALUES (?, ?, CURRENT_TIMESTAMP)
                    ON CONFLICT(key) DO UPDATE SET value = excluded.value, updated_at = CURRENT_TIMESTAMP
                """, (key, str(value)))

    def batch_update_person_status(self, ids, status):
        """批量更新人员启用/禁用状态"""
        if not ids:
            return
        with self.get_connection() as conn:
            conn.execute(
                f"UPDATE persons SET status = ?, updated_at = CURRENT_TIMESTAMP WHERE id IN ({','.join('?' * len(ids))})",
                [status] + list(ids)
            )

    # ==================== 操作日志 ====================

    def add_log(self, action, detail="", operator="system", ip_address=""):
        """添加操作日志"""
        with self.get_connection() as conn:
            conn.execute("""
                INSERT INTO operation_logs (operator, action, detail, ip_address)
                VALUES (?, ?, ?, ?)
            """, (operator, action, detail, ip_address))

    def get_logs(self, page=1, per_page=50, action=""):
        """获取操作日志"""
        offset = (page - 1) * per_page
        conditions = []
        params = []

        if action:
            conditions.append("action LIKE ?")
            params.append(f"%{action}%")

        where_clause = " AND ".join(conditions) if conditions else "1=1"

        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                f"SELECT * FROM operation_logs WHERE {where_clause} ORDER BY created_at DESC LIMIT ? OFFSET ?",
                params + [per_page, offset]
            )
            logs = [dict(row) for row in cursor.fetchall()]

            cursor.execute(
                f"SELECT COUNT(*) as total FROM operation_logs WHERE {where_clause}",
                params
            )
            total = cursor.fetchone()['total']

        return logs, total

    # ==================== 环境管理 ====================

    def add_environment(self, name, description="", work_start_hour=9, work_start_minute=0,
                       work_end_hour=18, work_end_minute=0, late_grace_minutes=15,
                       sign_in_required=1, sign_out_required=1, sign_mode="auto",
                       recognition_threshold=0.55, confirm_frames=3, sign_cooldown_seconds=60,
                       is_active=1, default_env=0):
        """添加环境"""
        with self.get_connection() as conn:
            cursor = conn.cursor()

            # 如果设置为默认环境，先取消其他默认环境
            if default_env:
                cursor.execute("UPDATE environments SET default_env = 0")

            cursor.execute("""
                INSERT INTO environments (name, description, work_start_hour, work_start_minute,
                                        work_end_hour, work_end_minute, late_grace_minutes,
                                        sign_in_required, sign_out_required, sign_mode,
                                        recognition_threshold, confirm_frames, sign_cooldown_seconds,
                                        is_active, default_env)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (name, description, work_start_hour, work_start_minute, work_end_hour,
                  work_end_minute, late_grace_minutes, sign_in_required, sign_out_required,
                  sign_mode, recognition_threshold, confirm_frames, sign_cooldown_seconds,
                  is_active, default_env))
            return cursor.lastrowid

    def update_environment(self, env_id, **kwargs):
        """更新环境信息"""
        allowed_fields = ['name', 'description', 'work_start_hour', 'work_start_minute',
                         'work_end_hour', 'work_end_minute', 'late_grace_minutes',
                         'sign_in_required', 'sign_out_required', 'sign_mode',
                         'recognition_threshold', 'confirm_frames', 'sign_cooldown_seconds',
                         'is_active', 'default_env', 'max_sign_count',
                         'sound_enabled', 'sound_volume', 'sound_text', 'sound_read_name']
        updates = []
        values = []

        for field, value in kwargs.items():
            if field in allowed_fields:
                updates.append(f"{field} = ?")
                values.append(value)

        if not updates:
            return False

        # 如果设置为默认环境，先取消其他默认环境
        if kwargs.get('default_env'):
            with self.get_connection() as conn:
                conn.execute("UPDATE environments SET default_env = 0")

        updates.append("updated_at = CURRENT_TIMESTAMP")
        values.append(env_id)

        with self.get_connection() as conn:
            conn.execute(
                f"UPDATE environments SET {', '.join(updates)} WHERE id = ?",
                values
            )
            return True

    def delete_environment(self, env_id):
        """删除环境"""
        with self.get_connection() as conn:
            conn.execute("DELETE FROM environments WHERE id = ?", (env_id,))
            return True

    def get_environment(self, env_id):
        """获取单个环境"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM environments WHERE id = ?", (env_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_all_environments(self, include_inactive=False):
        """获取所有环境"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            if include_inactive:
                cursor.execute("SELECT * FROM environments ORDER BY default_env DESC, name ASC")
            else:
                cursor.execute("SELECT * FROM environments WHERE is_active = 1 ORDER BY default_env DESC, name ASC")
            return [dict(row) for row in cursor.fetchall()]

    def get_active_environment(self):
        """获取当前激活的环境（默认环境）"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM environments WHERE default_env = 1 AND is_active = 1")
            row = cursor.fetchone()
            return dict(row) if row else None

    def set_default_environment(self, env_id):
        """设置默认环境"""
        with self.get_connection() as conn:
            conn.execute("UPDATE environments SET default_env = 0")
            conn.execute("UPDATE environments SET default_env = 1 WHERE id = ?", (env_id,))
            return True

    # ==================== 分类管理 ====================

    def add_category(self, name, parent_id=None, level=1, sort_order=0, description="", is_active=1):
        """添加分类"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO categories (name, parent_id, level, sort_order, description, is_active)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (name, parent_id, level, sort_order, description, is_active))
            return cursor.lastrowid

    def update_category(self, category_id, **kwargs):
        """更新分类信息"""
        allowed_fields = ['name', 'parent_id', 'level', 'sort_order', 'description', 'is_active']
        updates = []
        values = []

        for field, value in kwargs.items():
            if field in allowed_fields:
                updates.append(f"{field} = ?")
                values.append(value)

        if not updates:
            return False

        updates.append("updated_at = CURRENT_TIMESTAMP")
        values.append(category_id)

        with self.get_connection() as conn:
            conn.execute(
                f"UPDATE categories SET {', '.join(updates)} WHERE id = ?",
                values
            )
            return True

    def delete_category(self, category_id):
        """删除分类（级联删除子分类）"""
        with self.get_connection() as conn:
            conn.execute("DELETE FROM categories WHERE id = ?", (category_id,))
            return True

    def get_category(self, category_id):
        """获取单个分类"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM categories WHERE id = ?", (category_id,))
            row = cursor.fetchone()
            return dict(row) if row else None

    def get_all_categories(self, include_inactive=False, level=None):
        """获取所有分类"""
        with self.get_connection() as conn:
            cursor = conn.cursor()

            conditions = []
            params = []

            if not include_inactive:
                conditions.append("is_active = 1")

            if level is not None:
                conditions.append("level = ?")
                params.append(level)

            where_clause = " AND ".join(conditions) if conditions else "1=1"

            cursor.execute(
                f"SELECT * FROM categories WHERE {where_clause} ORDER BY level ASC, sort_order ASC",
                params
            )
            return [dict(row) for row in cursor.fetchall()]

    def get_categories_by_level(self, level, parent_id=None):
        """获取指定层级的分类"""
        with self.get_connection() as conn:
            cursor = conn.cursor()

            if parent_id is not None:
                cursor.execute(
                    "SELECT * FROM categories WHERE level = ? AND parent_id = ? AND is_active = 1 ORDER BY sort_order ASC",
                    (level, parent_id)
                )
            else:
                cursor.execute(
                    "SELECT * FROM categories WHERE level = ? AND is_active = 1 ORDER BY sort_order ASC",
                    (level,)
                )
            return [dict(row) for row in cursor.fetchall()]

    def get_category_tree(self):
        """获取完整的分类树"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM categories WHERE is_active = 1 ORDER BY level ASC, sort_order ASC")
            all_categories = [dict(row) for row in cursor.fetchall()]

            # 构建树形结构
            def build_tree(parent_id=None):
                return [
                    {**cat, 'children': build_tree(cat['id'])}
                    for cat in all_categories
                    if cat['parent_id'] == parent_id
                ]

            return build_tree(None)

    # ==================== 人脸审核管理 ====================

    def add_face_image(self, person_id, image_path, face_encoding=None, upload_source="mobile"):
        """添加人脸照片（待审核）"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO face_images (person_id, image_path, face_encoding, upload_source, approval_status)
                VALUES (?, ?, ?, ?, 'pending')
            """, (person_id, image_path, face_encoding, upload_source))
            return cursor.lastrowid

    def approve_face_image(self, face_image_id, approved_by):
        """批准人脸照片"""
        with self.get_connection() as conn:
            cursor = conn.cursor()

            # 获取人脸照片信息
            cursor.execute("SELECT person_id, face_encoding, image_path FROM face_images WHERE id = ?", (face_image_id,))
            face_data = cursor.fetchone()

            if not face_data:
                return False, "人脸照片不存在"

            # 更新审核状态
            cursor.execute("""
                UPDATE face_images
                SET approval_status = 'approved', approved_by = ?, approved_at = CURRENT_TIMESTAMP
                WHERE id = ?
            """, (approved_by, face_image_id))

            # 更新人员表的人脸编码（激活该人脸）
            cursor.execute("""
                UPDATE persons
                SET face_encoding = ?, face_image_path = ?
                WHERE id = ?
            """, (face_data['face_encoding'], face_data['image_path'], face_data['person_id']))

            # 将该人员的其他待审核/已批准的人脸设为非活跃
            cursor.execute("""
                UPDATE face_images
                SET is_active = 0
                WHERE person_id = ? AND id != ? AND approval_status = 'approved'
            """, (face_data['person_id'], face_image_id))

            conn.commit()
            return True, "审核通过"

    def reject_face_image(self, face_image_id, reject_reason=""):
        """拒绝人脸照片"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE face_images
                SET approval_status = 'rejected', reject_reason = ?, is_active = 0
                WHERE id = ?
            """, (reject_reason, face_image_id))
            return True

    def get_pending_faces(self, page=1, per_page=20):
        """获取待审核的人脸照片"""
        offset = (page - 1) * per_page
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT fi.*, p.name as person_name, p.employee_id
                FROM face_images fi
                JOIN persons p ON fi.person_id = p.id
                WHERE fi.approval_status = 'pending'
                ORDER BY fi.created_at ASC
                LIMIT ? OFFSET ?
            """, (per_page, offset))
            pending = [dict(row) for row in cursor.fetchall()]

            cursor.execute("SELECT COUNT(*) as total FROM face_images WHERE approval_status = 'pending'")
            total = cursor.fetchone()['total']

        return pending, total

    def get_person_face_images(self, person_id):
        """获取某人的所有人脸照片"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT fi.*,
                       CASE WHEN fi.approval_status = 'approved' AND fi.is_active = 1 THEN 1 ELSE 0 END as is_current
                FROM face_images fi
                WHERE fi.person_id = ?
                ORDER BY fi.created_at DESC
            """, (person_id,))
            return [dict(row) for row in cursor.fetchall()]

    # ==================== 人员环境关联 ====================

    def add_person_to_environment(self, person_id, environment_id, is_primary=0):
        """将人员添加到环境"""
        with self.get_connection() as conn:
            cursor = conn.cursor()

            # 如果设置为主要环境，先取消其他主要环境
            if is_primary:
                cursor.execute("UPDATE person_environment_rel SET is_primary = 0 WHERE person_id = ?", (person_id,))

            cursor.execute("""
                INSERT OR REPLACE INTO person_environment_rel (person_id, environment_id, is_primary)
                VALUES (?, ?, ?)
            """, (person_id, environment_id, is_primary))
            return True

    def remove_person_from_environment(self, person_id, environment_id):
        """从环境中移除人员"""
        with self.get_connection() as conn:
            conn.execute("DELETE FROM person_environment_rel WHERE person_id = ? AND environment_id = ?",
                        (person_id, environment_id))
            return True

    def get_person_environments(self, person_id):
        """获取人员所属的环境"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT e.*, per.is_primary, per.created_at as added_at
                FROM environments e
                JOIN person_environment_rel per ON e.id = per.environment_id
                WHERE per.person_id = ? AND e.is_active = 1
                ORDER BY per.is_primary DESC, e.name ASC
            """, (person_id,))
            return [dict(row) for row in cursor.fetchall()]

    def get_environment_persons(self, environment_id):
        """获取环境中的所有人员"""
        with self.get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT p.*, per.is_primary, per.created_at as added_to_env_at
                FROM persons p
                JOIN person_environment_rel per ON p.id = per.person_id
                WHERE per.environment_id = ? AND p.status = 1
                ORDER BY per.is_primary DESC, p.name ASC
            """, (environment_id,))
            return [dict(row) for row in cursor.fetchall()]

    # ==================== 数据导出 ====================

    def export_attendance_excel(self, start_date, end_date, filepath):
        """导出签到记录为Excel"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            records, _ = self.get_attendance_by_date_range(start_date, end_date, per_page=10000)

            if not records:
                return False, "没有数据可导出"

            df_data = []
            for r in records:
                df_data.append({
                    "姓名": r.get('name', ''),
                    "工号": r.get('employee_id', ''),
                    "部门": r.get('department', ''),
                    "职位": r.get('position', ''),
                    "签到类型": "签到" if r['sign_type'] == 'in' else "签退",
                    "签到时间": r['sign_time'],
                    "置信度": f"{r.get('confidence', 0):.2%}" if r.get('confidence') else '',
                    "备注": r.get('remark', '')
                })

            df = pd.DataFrame(df_data)

            wb = Workbook()
            ws = wb.active
            ws.title = "签到记录"

            # 表头样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 写入表头
            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            # 写入数据
            for row_idx, row_data in enumerate(df.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border

            # 自动调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 4, 30)

            wb.save(filepath)
            return True, f"成功导出 {len(df_data)} 条记录"

        except ImportError:
            return False, "缺少pandas或openpyxl库，请安装: pip install pandas openpyxl"
        except Exception as e:
            return False, f"导出失败: {str(e)}"

    def export_persons_excel(self, filepath):
        """导出人员列表为Excel"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            persons, _ = self.get_all_persons(include_inactive=True, per_page=10000)

            if not persons:
                return False, "没有数据可导出"

            df_data = []
            for p in persons:
                df_data.append({
                    "姓名": p.get('name', ''),
                    "工号": p.get('employee_id', ''),
                    "部门": p.get('department', ''),
                    "职位": p.get('position', ''),
                    "电话": p.get('phone', ''),
                    "邮箱": p.get('email', ''),
                    "状态": "启用" if p.get('status', 1) == 1 else "禁用",
                    "创建时间": p.get('created_at', ''),
                    "备注": p.get('remark', '')
                })

            df = pd.DataFrame(df_data)

            wb = Workbook()
            ws = wb.active
            ws.title = "人员列表"

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for col_idx, col_name in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

            for row_idx, row_data in enumerate(df.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border

            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 4, 30)

            wb.save(filepath)
            return True, f"成功导出 {len(df_data)} 条记录"

        except ImportError:
            return False, "缺少pandas或openpyxl库"
        except Exception as e:
            return False, f"导出失败: {str(e)}"

    # ==================== 设备配置相关 ====================

    def get_device_config(self):
        """获取设备配置"""
        return {
            'device_name': self.get_setting('device_name', ''),
            'device_id': self.get_setting('device_id', ''),
            'location': self.get_setting('location', ''),
            'description': self.get_setting('device_description', '')
        }

    def update_device_config(self, device_name=None, location=None, description=None):
        """更新设备配置"""
        if device_name:
            self.update_setting('device_name', device_name)
        if location:
            self.update_setting('location', location)
        if description:
            self.update_setting('device_description', description)

    def ensure_device_id(self):
        """确保设备ID存在"""
        device_id = self.get_setting('device_id')
        if not device_id:
            from device_discovery import DeviceDiscovery
            device_id = DeviceDiscovery.get_device_id()
            self.update_setting('device_id', device_id)
        return device_id


# 全局数据库实例
db = Database()
